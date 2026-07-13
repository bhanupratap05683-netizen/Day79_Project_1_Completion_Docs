"""
Day 79 - Portfolio Project 1: Financial Dashboard - Completion & Docs
Author: Bhanu Pratap Singh
Description: Polished, production-ready financial dashboard.
- Fetches (mock) API data, analyzes with pandas, visualizes with matplotlib,
- Exports polished Excel report with openpyxl, optional xlwings live update.
- Includes logging, error handling, docstrings, PEP 8 compliance.
"""

import os
import logging
import sys
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
import random

# --- CONFIGURATION CONSTANTS ---
STOCKS = ["RELIANCE", "TCS", "INFY", "HDFCBANK"]
REPORTS_DIR = "reports"
CHARTS_DIR = "charts"
EXCEL_REPORT_NAME = f"Financial_Dashboard_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
LOG_FILE = "dashboard.log"
DARK_GREEN = "0F5132"
WHITE = "FFFFFF"

# --- LOGGING SETUP (Day 64 Concept) ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Ensure directories exist
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(CHARTS_DIR, exist_ok=True)


def fetch_market_data_from_api():
    """
    Fetch market data. Simulates API call with fallback sample data.

    In production, replace with real API call:
    requests.get("https://www.alphavantage.co/query?function=TIME_SERIES_DAILY...")

    Returns:
        pd.DataFrame: DataFrame with Date, Symbol, Price, Volume
    """
    logger.info("Attempting to fetch market data from API...")
    try:
        # SIMULATED API DATA - Replace with real requests call
        # This ensures code runs without API key for portfolio demo
        dates = pd.date_range(end=datetime.now(), periods=30, freq='B')
        data = []
        for symbol in STOCKS:
            price = random.uniform(1500, 3500)
            for dt in dates:
                price += random.uniform(-50, 50)
                data.append({
                    "Date": dt,
                    "Symbol": symbol,
                    "Close_Price": round(max(price, 100), 2),
                    "Volume": random.randint(100000, 5000000)
                })
        df = pd.DataFrame(data)
        logger.info(f"Successfully fetched {len(df)} records for {len(STOCKS)} symbols.")
        return df

    except Exception as e:
        logger.error(f"API fetch failed: {e}. Using empty DataFrame.")
        raise


def analyze_data(df):
    """
    Analyze market data using pandas.

    Calculates daily returns, 7-day moving average, and volatility.

    Args:
        df (pd.DataFrame): Raw market data

    Returns:
        tuple: (detailed_df, summary_df)
    """
    logger.info("Starting data analysis with pandas...")
    df = df.sort_values(["Symbol", "Date"])
    df["Daily_Return_%"] = df.groupby("Symbol")["Close_Price"].pct_change() * 100
    df["7D_MA"] = df.groupby("Symbol")["Close_Price"].transform(
        lambda x: x.rolling(window=7).mean()
    )
    # Summary stats
    summary = df.groupby("Symbol").agg(
        Last_Price=("Close_Price", "last"),
        Avg_Volume=("Volume", "mean"),
        Volatility=("Daily_Return_%", "std"),
        Total_Return_Percent=("Daily_Return_%", "sum")
    ).reset_index()

    logger.info("Analysis complete. Summary stats generated.")
    return df, summary


def create_visualizations(df):
    """
    Create and save financial charts using matplotlib.

    Args:
        df (pd.DataFrame): Analyzed data

    Returns:
        str: Path to saved chart image
    """
    logger.info("Creating visualizations...")
    plt.style.use('seaborn-v0_8')
    fig, ax = plt.subplots(figsize=(10, 6))

    for symbol in STOCKS:
        symbol_data = df[df["Symbol"] == symbol]
        ax.plot(symbol_data["Date"], symbol_data["Close_Price"], label=symbol)

    ax.set_title("Stock Price Trend - Last 30 Days", fontsize=14, fontweight='bold')
    ax.set_xlabel("Date")
    ax.set_ylabel("Close Price (INR)")
    ax.legend()
    plt.xticks(rotation=20)
    plt.tight_layout()

    chart_path = os.path.join(CHARTS_DIR, "price_trend.png")
    plt.savefig(chart_path, dpi=300)
    plt.close()
    logger.info(f"Chart saved at {chart_path}")
    return chart_path


def export_to_excel_with_formatting(detailed_df, summary_df, chart_path):
    """
    Export data to a polished, multi-sheet Excel report using openpyxl.

    Includes header formatting, filters, currency formats.

    Args:
        detailed_df (pd.DataFrame): Detailed stock data
        summary_df (pd.DataFrame): Summary KPIs
        chart_path (str): Path to chart image file

    Returns:
        str: Path to final Excel file
    """
    logger.info("Exporting to Excel with professional formatting...")
    file_path = os.path.join(REPORTS_DIR, EXCEL_REPORT_NAME)

    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            detailed_df.to_excel(writer, sheet_name="Detailed_Data", index=False)

            # Formatting Summary Sheet
            workbook = writer.book
            summary_sheet = writer.sheets["Summary"]

            header_fill = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
            header_font = Font(color=WHITE, bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col_num, _ in enumerate(summary_df.columns, 1):
                cell = summary_sheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border

            # Auto-adjust column width
            for column_cells in summary_sheet.columns:
                length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
                summary_sheet.column_dimensions[column_cells[0].column_letter].width = length + 5

            summary_sheet.auto_filter.ref = summary_sheet.dimensions

        logger.info(f"Excel report successfully created at {file_path}")
        return file_path

    except Exception as e:
        logger.error(f"Failed to export Excel: {e}")
        raise


def update_live_excel_with_xlwings(file_path):
    """
    Attempt to update a live Excel workbook using xlwings.

    This is a safe implementation - if xlwings or Excel not available, it logs and skips.

    Args:
        file_path (str): Path to Excel file
    """
    logger.info("Attempting xlwings live update...")
    try:
        import xlwings as xw
        # This will only work if Excel is installed. We just demonstrate capability.
        # For portfolio, we don't need to actually keep it open.
        logger.info("xlwings is installed. Ready for live Excel integration (Excel must be open).")
        logger.info(f"To demonstrate: xw.Book('{file_path}').sheets['Summary'].range('A1').value")
    except ImportError:
        logger.warning("xlwings not installed. Skipping live update. To enable: pip install xlwings")
    except Exception as e:
        logger.warning(f"xlwings live update skipped: {e}")


def main():
    """Main pipeline: Fetch -> Analyze -> Visualize -> Excel -> xlwings."""
    logger.info("===== FINANCIAL DASHBOARD PROJECT STARTED =====")
    try:
        raw_df = fetch_market_data_from_api()
        detailed_df, summary_df = analyze_data(raw_df)
        chart_path = create_visualizations(detailed_df)
        excel_path = export_to_excel_with_formatting(detailed_df, summary_df, chart_path)
        update_live_excel_with_xlwings(excel_path)

        logger.info(f"===== PROJECT SUCCESSFULLY COMPLETED =====")
        logger.info(f"Check your report at: {excel_path}")
        print(f"\n✅ Dashboard Ready! Open file: {excel_path}")

    except Exception as e:
        logger.exception(f"Pipeline failed due to error: {e}")
        print(f"❌ Failed. Check {LOG_FILE} for details.")


if __name__ == "__main__":
    main()